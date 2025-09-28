import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException

# --- 1. CONFIGURATION ---

# IMPORTANT: Paste the "Profile Path" you copied from chrome://version here.
# To avoid errors, make sure to put an 'r' before the quotes, which marks it as a raw string.
# Example for Windows: r"C:\Users\YourUser\AppData\Local\Google\Chrome\User Data\Profile 2"
CHROME_PROFILE_PATH = r"C:\Users\pramo\AppData\Local\Google\Chrome\User Data\Profile 25"

# **NEW**: Add your Google email address here. It must match the account in your Chrome Profile.
# CORRECTED: The email address has been fixed.
GOOGLE_EMAIL = "rk91771432@gmail.com" # <--- IMPORTANT: UPDATE THIS IF NEEDED

# The name of the Excel file where your articles are stored.
ARTICLES_FILE = "leenusindia.xlsx"
# The specific sheet name within the Excel file to read articles from.
ARTICLES_SHEET_NAME = "Articles"
# The name of the Excel file where published links are logged.
OUTPUT_FILE = "published_articles_log.xlsx"

# --- 2. SETUP HELPER FUNCTIONS ---

def setup_output_file():
    """Creates the output log file with headers if it doesn't exist."""
    if not os.path.exists(OUTPUT_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Published URLs"
        sheet["A1"] = "Website"
        sheet["B1"] = "Article Title"
        sheet["C1"] = "Published URL"
        sheet["D1"] = "Timestamp"
        workbook.save(OUTPUT_FILE)

def find_next_article():
    """Reads the specified sheet and returns the details of the first un-published article."""
    try:
        workbook = openpyxl.load_workbook(ARTICLES_FILE)
        # Load the specific sheet named "Articles"
        sheet = workbook[ARTICLES_SHEET_NAME]
        # Assumes Column A=Title, B=Content, C=Status
        for row in range(2, sheet.max_row + 1): # Start from row 2 to skip header
            status = sheet[f"C{row}"].value
            if status is None or "published" not in str(status).lower():
                article_title = sheet[f"A{row}"].value
                article_content = sheet[f"B{row}"].value
                # Ensure title and content are not empty
                if article_title and article_content:
                    return row, article_title, article_content
        return None, None, None # No unpublished articles found
    except FileNotFoundError:
        print(f"ERROR: The file '{ARTICLES_FILE}' was not found.")
        return None, None, None
    except KeyError:
        print(f"ERROR: A sheet named '{ARTICLES_SHEET_NAME}' was not found in '{ARTICLES_FILE}'.")
        return None, None, None


def update_article_status(row_number, published_url):
    """Updates the status and adds the published URL to the 'Articles' sheet."""
    try:
        workbook = openpyxl.load_workbook(ARTICLES_FILE)
        sheet = workbook[ARTICLES_SHEET_NAME]
        sheet[f"C{row_number}"] = "Published"
        sheet[f"D{row_number}"] = published_url # Assumes Column D is for the URL
        workbook.save(ARTICLES_FILE)
    except Exception as e:
        print(f"Could not update the Excel file. Error: {e}")
    
def log_published_url(title, url):
    """Logs the successful post to the output file."""
    workbook = openpyxl.load_workbook(OUTPUT_FILE)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    sheet[f"A{next_row}"] = "Medium"
    sheet[f"B{next_row}"] = title
    sheet[f"C{next_row}"] = url
    sheet[f"D{next_row}"] = time.strftime("%Y-%m-%d %H:%M:%S")
    workbook.save(OUTPUT_FILE)

# --- 3. CORE AUTOMATION LOGIC ---
def post_to_medium(article_title, article_content):
    """Launches browser, logs in, posts the article, and returns the URL."""
    if "PASTE_YOUR_CHROME_PROFILE_PATH_HERE" in CHROME_PROFILE_PATH or "your_google_email@gmail.com" in GOOGLE_EMAIL:
        print("\nERROR: Please update the CHROME_PROFILE_PATH and GOOGLE_EMAIL variables in the script first.")
        return None

    options = Options()
    options.add_argument(f"user-data-dir={CHROME_PROFILE_PATH}")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    wait = WebDriverWait(driver, 40)
    published_url = None

    try:
        # --- REVISED LOGIN FLOW ---
        driver.get("https://medium.com/")
        print("Navigated to Medium homepage.")
        time.sleep(3) 

        print("Starting sign-in process...")
        
        print("Step 1: Looking for the 'Get started' button...")
        # Medium now uses a "Get started" button which opens the sign-in modal
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[3]/div[1]/div[1]/div/div/div/div[3]/div[5]/span/a/button'))).click()
        print("Step 1: Clicked 'Get started'.")
        time.sleep(2)

        print("Step 2: Looking for the 'Sign in' link in the pop-up...")
        # **FIXED**: Using a more stable selector based on the link's URL (href)
        signin_link_selector = "/html/body/div[4]/div[2]/div/div/div/div[1]/div[1]/div[1]/p/button"
        wait.until(EC.element_to_be_clickable((By.XPATH, signin_link_selector))).click()
        print("Step 2: Clicked 'Sign in' link.")
        time.sleep(2)

        print("Step 3: Looking for 'Sign in with Google' button...")
        google_button_selector = "//button[contains(., 'Sign in with Google')]"
        wait.until(EC.element_to_be_clickable((By.XPATH, google_button_selector))).click()
        print("Step 3: Clicked 'Sign in with Google'.")
        
        print("Step 4: Switching to Google login window...")
        original_window = driver.current_window_handle
        wait.until(EC.number_of_windows_to_be(2))
        for window_handle in driver.window_handles:
            if window_handle != original_window:
                driver.switch_to.window(window_handle)
                break
        
        print(f"Step 5: Looking for account '{GOOGLE_EMAIL}' to click...")
        account_button = wait.until(EC.element_to_be_clickable((By.XPATH, f"//div[text()='{GOOGLE_EMAIL}']")))
        account_button.click()
        print("Step 5: Clicked account.")
        
        driver.switch_to.window(original_window)
        # Add a wait to ensure the main page has time to reflect the login status
        wait.until(EC.presence_of_element_located((By.XPATH, "//a[text()='Write']")))
        print("Step 6: Switched back to main window. Login successful.")

        # --- ARTICLE POSTING ---
        driver.get("https://medium.com/new-story")
        print("Navigated to the new story page.")

        time.sleep(5)
        title_field = wait.until(EC.visibility_of_element_located((By.XPATH, "//textarea[@aria-label='Title']")))
        title_field.send_keys(article_title)
        print("Article title entered.")
        
        time.sleep(2) 
        title_field.click()
        title_field.send_keys(Keys.ENTER)
        
        time.sleep(2) 
        content_field = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@data-placeholder='Tell your story…']")))
        content_field.send_keys(article_content)
        print("Article content entered.")

        time.sleep(3)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Publish']"))).click()
        print("Clicked 'Publish' menu button.")

        time.sleep(2)
        final_publish_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='publish-button']")))
        final_publish_button.click()
        print("Clicked final 'Publish now'.")

        time.sleep(5)
        wait.until(EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Story has been published')]")))
        published_url = driver.current_url.split('?')[0] # Clean the URL
        print(f"Successfully published! URL: {published_url}")

    except Exception as e:
        print(f"\nAn error occurred: {e}")
        driver.save_screenshot("error_screenshot.png")
        print("Saved a screenshot to 'error_screenshot.png' for debugging.")
    finally:
        print("Closing browser in 5 seconds...")
        time.sleep(5)
        driver.quit()
        return published_url

# --- 4. MAIN EXECUTION BLOCK ---
if __name__ == "__main__":
    setup_output_file()
    
    print(f"Searching for an article to publish in '{ARTICLES_FILE}' (Sheet: '{ARTICLES_SHEET_NAME}')...")
    row_num, title, content = find_next_article()
    
    if title and content:
        print(f"Found article to post: '{title}'")
        final_url = post_to_medium(title, content)
        
        if final_url:
            update_article_status(row_num, final_url)
            log_published_url(title, final_url)
            print(f"\nProcess complete. Article status updated in {ARTICLES_FILE}.")
        else:
            print("\nProcess failed. Could not retrieve the published URL.")
    else:
        print("No new articles to publish were found.")

